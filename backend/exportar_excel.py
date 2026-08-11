"""
Exportação Excel dos relatórios do motor de análise — extraído de
analisador-monitoria-2d/app.py (mesmas funções, sem nada de Tkinter).
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ImagemExcel

from engine.recursos import CAMINHO_LOGO, CAMINHO_LOGO_ICONE, NOME_SISTEMA, NOME_EMPRESA

# Paleta da marca, igual à do app (dashboard/src/index.css :root):
# preto quente no cabeçalho, Soft Fawn dourado como acento.
# O corpo da planilha fica claro de propósito — planilha é feita para ser
# filtrada, editada e impressa; fundo escuro em 650 mil células atrapalha os
# três usos e briga com qualquer tema do Excel do destinatário.
COR_CABECALHO = "25252C"
COR_CABECALHO_TEXTO = "DABB6C"
COR_FAIXA_TOTAIS = "F6EEDA"
COR_ZEBRA = "F7F5F0"
COR_POSITIVO = "2E7D52"
COR_NEGATIVO = "B4322A"

#: Variação com sinal (positivo é bom) e perda registrada como positivo
#: (positivo é ruim). Espelha COLUNAS_DELTA / COLUNAS_PERDA de exportar_html.py
#: e de ResultTable.tsx.
COLUNAS_DELTA = {
    "Desempenho_Pct", "Ganho_Perda", "Diferenca_Receita", "Variacao_Percentual",
    "Variacao_Global_Periodo_Pct", "Tendencia_Pct",
}
COLUNAS_PERDA = {
    "Reducao_Receita", "Reducao_Percentual", "Renuncia", "Renuncia_Acumulada",
    "Renuncia_Percentual", "Perda_Receita", "Receita_Sob_Risco",
    "Impacto_Financeiro_Churn", "Maior_Retracao_Individual_Pct",
}

# Catálogo de relatórios prontos, agrupados por categoria. Cada item mapeia um
# título de negócio para as chaves internas calculadas por
# analise_funil.gerar_analises_completas.
CATALOGO_RELATORIOS = [
    ("Vendas", [
        ("top_fabricantes", "Venda por Fabricante (Top Fabricantes)"),
        ("top_produtos", "Venda por Produto (Top Produtos)"),
        ("comparativo_receita", "Comparativo de Receita x Ano Anterior (por produto)"),
    ]),
    ("Segmentação e Poder de Compra", [
        ("abc", "Concentrado / Faturamento e Segmentação de Clientes (ABC)"),
        ("poder_compra_clientes", "Poder de Compra por Cliente (3 maiores meses)"),
        ("migracao_abc", "Migração de Grupo (inclui resumo e score por cliente)"),
    ]),
    ("Tendências e Alertas", [
        ("evolucao_produtos", "Tendência de Produtos"),
        ("alertas_queda", "Alertas de Queda Consecutiva"),
        ("erosao_clientes", "Erosão de Clientes por Produto"),
        ("sem_venda", "Sem Venda"),
    ]),
    ("Boletins", [
        ("produtos_em_alta", "Boletim: Produtos em Alta"),
        ("produtos_em_queda", "Boletim: Produtos em Queda"),
        ("clientes_queda_qtd", "Boletim: Clientes em Queda de Quantidade"),
        ("correlacao_produto_cliente", "Boletim: Correlação Produto x Cliente"),
        ("impacto_financeiro_churn", "Boletim: Impacto Financeiro do Churn"),
    ]),
    ("Alvos", [
        ("mais_atacado", "Mais Atacado"),
        ("liquidez", "Liquidez (Estoque + Vendas)"),
    ]),
]

NOMES_ANALISE = {
    "top_produtos": "Top_Produtos",
    "top_fabricantes": "Top_Fabricantes",
    "comparativo_receita": "Comparativo_Receita",
    "poder_compra_clientes": "Poder_Compra_Clientes",
    "evolucao_produtos": "Evolucao_Produtos",
    "alertas_queda": "Alertas_Queda",
    "erosao_clientes": "Erosao_Clientes",
    "sem_venda": "Sem_Venda",
    "abc": "ABC_Clientes",
    "abc_produtos": "ABC_Produtos",
    "migracao_abc": "Migracao_ABC",
    "migracao_resumo": "Migracao_Resumo",
    "migracao_score_clientes": "Migracao_Score_Clientes",
    "produtos_em_alta": "Produtos_Em_Alta",
    "produtos_em_queda": "Produtos_Em_Queda",
    "clientes_queda_qtd": "Clientes_Queda_Qtd",
    "correlacao_produto_cliente": "Correlacao_Prod_Cliente",
    "impacto_financeiro_churn": "Impacto_Financeiro_Churn",
    "mais_atacado": "Mais_Atacado",
    "liquidez_estoque": "Liquidez_Estoque",
    "liquidez_vendas": "Liquidez_Vendas",
}

# Análises que ganham uma faixa "Totais" acima do cabeçalho, com a soma destas
# colunas. Espelhado em dashboard/src/components/analisador/ResultTable.tsx,
# que monta o mesmo resumo acima da tabela na tela.
COLUNAS_TOTALIZAVEIS_POR_ANALISE = {
    "comparativo_receita": ["Receita_Ano_Anterior", "Receita_Ano_Atual", "Ganho_Perda"],
}

COLUNAS_MOEDA_POR_ANALISE = {
    "top_produtos": ["Receita"],
    "top_fabricantes": ["Receita"],
    "comparativo_receita": ["Receita_Ano_Anterior", "Receita_Ano_Atual", "Ganho_Perda"],
    "poder_compra_clientes": ["Poder_De_Compra"],
    "evolucao_produtos": ["Receita", "Receita_Periodo_Anterior"],
    "alertas_queda": ["Receita_Ultimo_Periodo", "Receita_Primeiro_Periodo"],
    "erosao_clientes": ["Receita", "Receita_Periodo_Anterior", "Reducao_Receita"],
    "sem_venda": ["Receita", "Receita_Periodo_Anterior", "Reducao_Receita"],
    "abc": ["Receita", "Renuncia", "Renuncia_Acumulada"],
    "abc_produtos": ["Receita", "Renuncia", "Renuncia_Acumulada"],
    "migracao_abc": [],
    "migracao_resumo": [],
    "migracao_score_clientes": [],
    "produtos_em_alta": ["Receita_Periodo_Anterior", "Receita_Periodo_Atual", "Diferenca_Receita", "Total_Ano_Atual"],
    "produtos_em_queda": ["Receita_Periodo_Anterior", "Receita_Periodo_Atual", "Diferenca_Receita", "Total_Ano_Atual"],
    "clientes_queda_qtd": ["Perda_Receita"],
    "correlacao_produto_cliente": ["Reducao_Receita"],
    "impacto_financeiro_churn": ["Receita_Sob_Risco"],
    "mais_atacado": ["Receita Acumulada 11 Meses"],
    "liquidez_estoque": ["Preço_médio_de_venda", "Preço_médio_cmv", "Último_custo"],
    "liquidez_vendas": [],
}


def calcular_faixa_totais(chave_analise, df):
    """Totais da análise, quando ela tem faixa própria — senão None.

    Soma as colunas declaradas em COLUNAS_TOTALIZAVEIS_POR_ANALISE (o relatório
    traz todos os itens, sem top N, então a soma das linhas É o total) e deriva
    o desempenho a partir das duas receitas.
    """
    colunas = COLUNAS_TOTALIZAVEIS_POR_ANALISE.get(chave_analise)
    if not colunas or df is None or df.empty:
        return None

    faixa = {}
    primeira = df.columns[0]
    faixa[primeira] = "Totais"
    for coluna in colunas:
        if coluna in df.columns:
            faixa[coluna] = float(pd.to_numeric(df[coluna], errors="coerce").fillna(0).sum())

    anterior = faixa.get("Receita_Ano_Anterior")
    atual = faixa.get("Receita_Ano_Atual")
    if anterior and atual is not None and "Desempenho_Pct" in df.columns:
        faixa["Desempenho_Pct"] = (atual - anterior) / anterior * 100
    return faixa


def _eh_coluna_percentual(nome_coluna):
    """Colunas em escala 0–100 do motor (Reducao_Percentual, Tendencia_Pct...)."""
    return "Percentual" in str(nome_coluna) or str(nome_coluna).endswith("_Pct")


def _ajustar_largura_colunas(planilha, max_linhas=2000):
    """Estima larguras sem reler planilhas gigantes por completo."""
    ultima_linha_amostra = min(planilha.max_row, max_linhas)
    for indice_coluna in range(1, planilha.max_column + 1):
        maior_comprimento = 0
        letra_coluna = get_column_letter(indice_coluna)
        # Não usar ``planilha.columns``: o openpyxl materializa a coluna
        # inteira antes do slice, consumindo tempo e memória em relatórios grandes.
        for indice_linha in range(1, ultima_linha_amostra + 1):
            celula = planilha.cell(row=indice_linha, column=indice_coluna)
            valor = str(celula.value) if celula.value is not None else ""
            maior_comprimento = max(maior_comprimento, len(valor))
        planilha.column_dimensions[letra_coluna].width = min(maior_comprimento + 2, 45)


def _formatar_cabecalho(planilha, linha=1):
    preenchimento = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")
    fonte = Font(color=COR_CABECALHO_TEXTO, bold=True)
    borda_accent = Border(bottom=Side(style="medium", color=COR_CABECALHO_TEXTO))
    for celula in planilha[linha]:
        celula.fill = preenchimento
        celula.font = fonte
        celula.border = borda_accent
        celula.alignment = Alignment(horizontal="center", vertical="center")
    planilha.row_dimensions[linha].height = 22


def _formatar_faixa_totais(planilha, qtd_colunas):
    """Destaca a linha 1 (faixa de totais), que fica acima do cabeçalho."""
    preenchimento = PatternFill(start_color=COR_FAIXA_TOTAIS, end_color=COR_FAIXA_TOTAIS, fill_type="solid")
    for indice in range(1, qtd_colunas + 1):
        celula = planilha.cell(row=1, column=indice)
        celula.fill = preenchimento
        celula.font = Font(bold=True)


#: Acima disso não pinta zebra nem colore variação célula por célula. Cada
#: célula estilizada é um objeto a mais no XML da planilha: em relatório de
#: centenas de milhares de linhas o ganho visual não paga o tempo de geração nem
#: o tamanho do arquivo. Os relatórios do catálogo ficam muito abaixo do limite.
MAX_LINHAS_ESTILIZADAS = 5000


def _pintar_zebra(planilha, linha_cabecalho, linhas_com_valor, qtd_colunas):
    """Linhas alternadas, como na tela."""
    if len(linhas_com_valor) > MAX_LINHAS_ESTILIZADAS:
        return
    preenchimento = PatternFill(start_color=COR_ZEBRA, end_color=COR_ZEBRA, fill_type="solid")
    for posicao, linha in enumerate(sorted(linhas_com_valor)):
        if posicao % 2 == 0:
            continue
        for coluna in range(1, qtd_colunas + 1):
            planilha.cell(row=linha, column=coluna).fill = preenchimento


def _colorir_deltas(planilha, colunas, linha_cabecalho):
    """Verde/vermelho nas colunas de variação e de perda — mesmo semáforo da tela.

    Sem seta: no Excel a seta teria que virar texto dentro da célula e quebraria
    o valor numérico (soma, filtro, gráfico). Cor resolve sem mexer no dado.
    """
    alvo = [
        (indice, nome) for indice, nome in enumerate(colunas, start=1)
        if nome in COLUNAS_DELTA or nome in COLUNAS_PERDA
    ]
    if not alvo or planilha.max_row - 1 > MAX_LINHAS_ESTILIZADAS:
        return
    for indice_coluna, nome in alvo:
        eh_perda = nome in COLUNAS_PERDA
        for linha in range(1, planilha.max_row + 1):
            if linha == linha_cabecalho:
                continue
            celula = planilha.cell(row=linha, column=indice_coluna)
            valor = celula.value
            if not isinstance(valor, (int, float)) or isinstance(valor, bool) or valor == 0:
                continue
            if eh_perda:
                cor = COR_NEGATIVO if valor > 0 else None
            else:
                cor = COR_POSITIVO if valor > 0 else COR_NEGATIVO
            if cor:
                celula.font = Font(color=cor, bold=True)


def _inserir_logo(planilha, coluna_ancora, linha_ancora=1, altura_pixels=34):
    """Insere a marca da 2D Consultores em miniatura, sem sobrepor os dados."""
    if not os.path.exists(CAMINHO_LOGO_ICONE):
        return
    try:
        imagem = ImagemExcel(CAMINHO_LOGO_ICONE)
        proporcao = imagem.width / imagem.height
        imagem.height = altura_pixels
        imagem.width = altura_pixels * proporcao
        planilha.add_image(imagem, f"{get_column_letter(coluna_ancora)}{linha_ancora}")
    except Exception:
        pass  # ausência da logo não deve impedir a geração do relatório


def _escrever_dataframe(workbook, nome_aba, df, colunas_moeda=None, faixa_totais=None):
    """Escreve o DataFrame na aba. `faixa_totais` vira uma linha acima do cabeçalho.

    A faixa fica FORA da tabela de propósito: um "Totais" como linha de dados
    entraria no autofiltro, se moveria ao ordenar e duplicaria qualquer soma
    feita sobre a coluna.
    """
    if nome_aba in workbook.sheetnames:
        planilha = workbook[nome_aba]
    else:
        planilha = workbook.create_sheet(nome_aba)

    colunas_moeda = colunas_moeda or []
    df_para_exportar = df.reset_index() if df.index.name or isinstance(df.index, pd.MultiIndex) else df

    linha_cabecalho = 1
    if faixa_totais:
        planilha.append([faixa_totais.get(coluna) for coluna in df_para_exportar.columns])
        linha_cabecalho = 2

    planilha.append(list(map(str, df_para_exportar.columns)))
    # itertuples evita criar uma Series por linha; diferença grande em bases
    # com centenas de milhares de registros.
    for linha in df_para_exportar.itertuples(index=False, name=None):
        planilha.append(linha)

    linhas_com_valor = [
        linha for linha in range(1, planilha.max_row + 1) if linha != linha_cabecalho
    ]
    for indice_coluna, nome_coluna in enumerate(df_para_exportar.columns, start=1):
        if nome_coluna in colunas_moeda:
            formato = 'R$ #,##0.00'
        elif _eh_coluna_percentual(nome_coluna):
            # Valores já vêm em escala 0–100 — o "%" é literal (não o formato
            # percentual do Excel, que multiplicaria por 100 de novo).
            formato = '0.00"%"'
        else:
            continue
        for linha in linhas_com_valor:
            planilha.cell(row=linha, column=indice_coluna).number_format = formato

    _pintar_zebra(planilha, linha_cabecalho, linhas_com_valor, len(df_para_exportar.columns))
    _colorir_deltas(planilha, df_para_exportar.columns, linha_cabecalho)

    if faixa_totais:
        _formatar_faixa_totais(planilha, len(df_para_exportar.columns))

    _formatar_cabecalho(planilha, linha_cabecalho)
    _ajustar_largura_colunas(planilha)
    _inserir_logo(planilha, coluna_ancora=len(df_para_exportar.columns) + 2)
    return planilha


def _criar_capa(workbook, resultados_analise, nome_usuario="", nome_empresa=""):
    """Primeira aba do relatório: logo, identidade da empresa e sumário do que foi gerado."""
    capa = workbook.create_sheet("Capa", 0)
    capa.sheet_view.showGridLines = False
    capa.column_dimensions["A"].width = 4
    capa.column_dimensions["B"].width = 60

    if os.path.exists(CAMINHO_LOGO):
        try:
            imagem = ImagemExcel(CAMINHO_LOGO)
            proporcao = imagem.width / imagem.height
            imagem.height = 130
            imagem.width = 130 * proporcao
            capa.add_image(imagem, "B2")
        except Exception:
            pass

    capa["B10"] = NOME_SISTEMA
    capa["B10"].font = Font(size=20, bold=True, color=COR_CABECALHO)
    capa["B11"] = NOME_EMPRESA
    capa["B11"].font = Font(size=12, color="666666")

    linha_info = 13
    if nome_empresa:
        capa[f"B{linha_info}"] = f"Empresa analisada: {nome_empresa}"
        capa[f"B{linha_info}"].font = Font(size=13, bold=True, color=COR_CABECALHO)
        linha_info += 1
    if nome_usuario:
        capa[f"B{linha_info}"] = f"Gerado por: {nome_usuario}"
        capa[f"B{linha_info}"].font = Font(size=10, italic=True, color="666666")
        linha_info += 1
    capa[f"B{linha_info}"] = f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    capa[f"B{linha_info}"].font = Font(size=10, italic=True, color="666666")

    linha = linha_info + 3
    capa[f"B{linha}"] = "Granularidades incluídas neste relatório:"
    capa[f"B{linha}"].font = Font(bold=True)
    for granularidade in resultados_analise.keys():
        linha += 1
        capa[f"B{linha}"] = f"•  {granularidade}"
    return capa


def exportar_relatorio_excel(caminho_saida, resultados_analise, relatorios_personalizados=None, nome_usuario="", nome_empresa=""):
    """
    Gera o arquivo .xlsx com uma aba por (análise x granularidade), formatado
    com cabeçalhos destacados, moeda BRL, largura de coluna automática e a
    logo da empresa em cada aba (mais uma capa de apresentação).
    """
    workbook = Workbook()
    workbook.remove(workbook.active)  # remove a aba padrão vazia
    _criar_capa(workbook, resultados_analise, nome_usuario, nome_empresa)

    for granularidade, analises in resultados_analise.items():
        for chave_analise, df_analise in analises.items():
            nome_base = NOMES_ANALISE.get(chave_analise, chave_analise)
            if granularidade == "Alvos":
                nome_aba = str(nome_base)[:31]
            else:
                nome_aba = f"{nome_base}_{granularidade}"[:31]  # limite do Excel
            if df_analise is None or df_analise.empty:
                planilha = workbook.create_sheet(nome_aba)
                planilha.append(["Sem dados para esta análise/granularidade."])
                continue
            _escrever_dataframe(
                workbook, nome_aba, df_analise,
                COLUNAS_MOEDA_POR_ANALISE.get(chave_analise),
                faixa_totais=calcular_faixa_totais(chave_analise, df_analise),
            )

    if relatorios_personalizados:
        for nome_relatorio, tabela in relatorios_personalizados.items():
            nome_aba = f"Custom_{nome_relatorio}"[:31]
            _escrever_dataframe(workbook, nome_aba, tabela)

    if len(workbook.sheetnames) <= 1:
        workbook.create_sheet("Sem_Dados")

    workbook.save(caminho_saida)
